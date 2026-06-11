"""Project specification parser and initial motor design generator.

Handles:
- Loading project specs from files (JSON, YAML, TXT)
- Parsing natural language spec text
- Generating initial electromagnetic design using analytical sizing equations
"""

from dataclasses import dataclass, field, asdict
from enum import Enum
from typing import Optional
import json
import os
import math
import re

from .knowledge import (
    MotorType, CoolingType, SLOT_POLE_RECOMMENDATIONS, PM_MATERIALS
)


# ---------------------------------------------------------------------------
# Project Specification
# ---------------------------------------------------------------------------

@dataclass
class ProjectSpec:
    """Complete motor design project specification."""
    project_name: str = "Unnamed Project"
    motor_type: str = "IPM"
    rated_power_kw: float = 0.0
    rated_speed_rpm: float = 0.0
    max_speed_rpm: float = 0.0
    rated_torque_nm: float = 0.0       # auto-calculated
    peak_torque_nm: float = 0.0
    dc_voltage_v: float = 0.0
    max_current_arms: float = 0.0
    target_efficiency_pct: float = 95.0
    max_outer_diameter_mm: float = 300.0
    max_stack_length_mm: float = 200.0
    cooling: str = "water_jacket"
    application: str = "industrial"
    phases: int = 3
    max_current_density: float = 15.0

    def __post_init__(self):
        if self.rated_torque_nm == 0 and self.rated_power_kw > 0 and self.rated_speed_rpm > 0:
            self.rated_torque_nm = 9550 * self.rated_power_kw / self.rated_speed_rpm


# ---------------------------------------------------------------------------
# Initial Design Generator
# ---------------------------------------------------------------------------

@dataclass
class InitialDesign:
    """Complete initial motor design parameters."""
    motor_type: str = "IPM"
    # Geometry
    stator_outer_diameter_mm: float = 200
    stator_bore_mm: float = 120
    airgap_mm: float = 1.0
    stack_length_mm: float = 150
    slot_number: int = 36
    pole_number: int = 6
    tooth_width_mm: float = 6
    slot_depth_mm: float = 25
    stator_yoke_width_mm: float = 15
    slot_opening_mm: float = 3.0
    # Rotor / PM
    magnet_thickness_mm: float = 5.0
    magnet_arc_deg: float = 140
    pole_embrace: float = 0.78
    rotor_outer_diameter_mm: float = 118
    # Winding
    turns_per_coil: int = 8
    strands_per_turn: int = 2
    wire_diameter_mm: float = 1.5
    winding_layers: int = 2
    # Materials
    lamination_material: str = "M250-35A"
    magnet_material: str = "NdFeB_42UH"
    # Estimated performance
    estimated_torque_nm: float = 0
    estimated_efficiency_pct: float = 0
    estimated_current_density: float = 0
    # Metadata
    design_notes: list = field(default_factory=list)

    def to_dict(self) -> dict:
        return asdict(self)

    def to_motorcad_params(self) -> dict:
        """Convert to Motor-CAD variable names for easy import."""
        return {
            "Stator_Lam_Dia": self.stator_outer_diameter_mm,
            "Stator_Bore": self.stator_bore_mm,
            "Airgap": self.airgap_mm,
            "Stack_Length": self.stack_length_mm,
            "Slot_Number": self.slot_number,
            "Pole_Number": self.pole_number,
            "Tooth_Width": self.tooth_width_mm,
            "Slot_Depth": self.slot_depth_mm,
            "Stator_Yoke_Width": self.stator_yoke_width_mm,
            "Slot_Opening": self.slot_opening_mm,
            "Magnet_Thickness": self.magnet_thickness_mm,
            "Magnet_Arc": self.magnet_arc_deg,
            "Pole_Embrace": self.pole_embrace,
        }


class InitialDesignGenerator:
    """Generate initial motor design from project specs using analytical sizing."""

    # Shear stress lookup by cooling type (kN/m^2)
    SHEAR_STRESS = {
        "natural_convection": (10, 20),
        "forced_air": (20, 35),
        "water_jacket": (35, 55),
        "oil_spray": (45, 65),
        "oil_immersion": (50, 70),
    }

    # D/L ratio by application
    DL_RATIO = {
        "traction": (0.8, 1.2),       # EV traction: compact
        "industrial": (0.5, 0.8),     # Industrial: longer
        "servo": (1.0, 1.5),          # Servo: large diameter
        "direct_drive": (1.2, 2.0),   # Direct drive: large diameter
    }

    def generate(self, spec: ProjectSpec) -> InitialDesign:
        """Generate complete initial motor design."""
        design = InitialDesign(motor_type=spec.motor_type)
        notes = []

        # Step 1: Get shear stress range
        shear_range = self.SHEAR_STRESS.get(spec.cooling, (20, 35))
        shear_stress = (shear_range[0] + shear_range[1]) / 2 * 1000  # Pa

        # Step 2: Calculate D^2 * L from torque
        # T = pi/2 * sigma * D^2 * L
        torque = spec.rated_torque_nm if spec.rated_torque_nm > 0 else 100
        D2L = 2 * torque / (math.pi * shear_stress)
        notes.append(f"D^2*L = {D2L:.6f} m^3 (shear stress = {shear_stress/1000:.0f} kPa)")

        # Step 3: Get D/L ratio range
        dl_range = self.DL_RATIO.get(spec.application, (0.6, 1.0))
        dl_ratio = (dl_range[0] + dl_range[1]) / 2

        # Step 4: Solve for D and L: D^3 * (L/D) = D2L * (L/D) ? no...
        # D2L = D^2 * L, and L = D / dl_ratio ? no, L/D = 1/dl_ratio...
        # D2L = D^2 * (D * dl_ratio) ? no...
        # D2L = D^2 * L, D/L = dl_ratio, so L = D/dl_ratio
        # D2L = D^2 * D/dl_ratio = D^3 / dl_ratio
        # D = (D2L * dl_ratio)^(1/3)
        D_m = (D2L * dl_ratio) ** (1/3)
        L_m = D_m / dl_ratio

        D_mm = D_m * 1000
        L_mm = L_m * 1000

        notes.append(f"Raw sizing: D={D_mm:.0f}mm, L={L_mm:.0f}mm")

        # Step 5: Apply constraints
        D_mm = min(D_mm, spec.max_outer_diameter_mm * 0.75)  # 75% for yoke+slots
        L_mm = min(L_mm, spec.max_stack_length_mm)
        notes.append(f"After constraints: D={D_mm:.0f}mm, L={L_mm:.0f}mm")

        design.stator_bore_mm = round(D_mm)
        design.stack_length_mm = round(L_mm)

        # Step 6: Select slot/pole combination
        slots, poles = self._select_slot_pole(spec)
        design.slot_number = slots
        design.pole_number = poles
        notes.append(f"Slot/pole: {slots}s/{poles}p")

        # Step 7: Size stator geometry
        # Airgap: scale with bore
        design.airgap_mm = round(max(0.5, D_mm * 0.005), 1)
        design.rotor_outer_diameter_mm = design.stator_bore_mm - 2 * design.airgap_mm

        # Stator yoke: size for ~1.4T at rated flux
        # Pole pitch = pi * D / (2p)
        pole_pitch = math.pi * D_mm / (2 * poles)
        design.stator_yoke_width_mm = round(pole_pitch * 0.35)
        design.stator_yoke_width_mm = max(8, design.stator_yoke_width_mm)

        # Stator outer diameter
        design.stator_outer_diameter_mm = design.stator_bore_mm + 2 * design.stator_yoke_width_mm + 60
        if design.stator_outer_diameter_mm > spec.max_outer_diameter_mm:
            design.stator_outer_diameter_mm = spec.max_outer_diameter_mm

        # Tooth width: size for ~1.6T
        # Flux per pole = B_gap * pole_pitch * L
        # Tooth flux = flux_per_pole / (slots/pole/phases * ...)
        B_gap = 0.85  # target airgap flux density
        tooth_pitch = math.pi * D_mm / slots
        design.tooth_width_mm = round(tooth_pitch * 0.45)
        design.tooth_width_mm = max(3, design.tooth_width_mm)

        # Slot depth
        design.slot_depth_mm = round(
            (design.stator_outer_diameter_mm - design.stator_bore_mm) / 2
            - design.stator_yoke_width_mm
        )
        design.slot_opening_mm = round(max(2.0, tooth_pitch * 0.15), 1)

        # Step 8: Magnet design
        design.magnet_thickness_mm = round(3 + D_mm * 0.02, 1)  # empirical
        design.magnet_thickness_mm = max(3, min(10, design.magnet_thickness_mm))
        design.pole_embrace = 0.78
        design.magnet_arc_deg = round(180 * design.pole_embrace / poles * poles)

        # Step 9: Winding estimation
        # EMF per phase ~ 4.44 * f * N * phi
        # For rated speed, estimate turns
        freq = spec.rated_speed_rpm * poles / 120 if spec.rated_speed_rpm > 0 else 200
        phi_per_pole = B_gap * (pole_pitch / 1000) * (L_mm / 1000)  # Wb
        back_emf_target = spec.dc_voltage_v / math.sqrt(3) * 0.85 if spec.dc_voltage_v > 0 else 200
        turns_per_phase = back_emf_target / (4.44 * freq * phi_per_pole * 0.92) if freq > 0 else 40
        coils_per_phase = slots / 3
        design.turns_per_coil = max(1, round(turns_per_phase / coils_per_phase))

        # Wire gauge: based on current density and slot area
        if spec.rated_power_kw > 0 and spec.dc_voltage_v > 0:
            rated_current = spec.rated_power_kw * 1000 / (spec.dc_voltage_v * 0.9)
        else:
            rated_current = 100
        wire_area = rated_current / spec.max_current_density / design.turns_per_coil
        design.wire_diameter_mm = round(2 * math.sqrt(wire_area / math.pi), 2)
        design.wire_diameter_mm = max(0.5, min(3.0, design.wire_diameter_mm))
        design.strands_per_turn = max(1, round(rated_current / (spec.max_current_density * math.pi * (design.wire_diameter_mm/2)**2) / design.turns_per_coil))

        design.estimated_current_density = round(spec.max_current_density, 1)

        # Step 10: Material selection
        design.lamination_material = self._recommend_lamination(spec)
        design.magnet_material = self._recommend_magnet(spec)

        # Performance estimates
        design.estimated_torque_nm = round(torque * 0.9, 1)  # 90% of ideal
        design.estimated_efficiency_pct = round(spec.target_efficiency_pct - 2.0, 1)  # first pass

        design.design_notes = notes
        return design

    def _select_slot_pole(self, spec: ProjectSpec) -> tuple:
        """Select optimal slot/pole combination."""
        speed = spec.rated_speed_rpm if spec.rated_speed_rpm > 0 else 3000
        candidates = []

        for (s, p), info in SLOT_POLE_RECOMMENDATIONS.items():
            # Filter by speed frequency limit
            freq = speed * p / 120
            if freq > 800:
                continue
            # Favor high winding factor
            score = info["kw"] * 100
            # Favor higher pole count for torque density
            score += p * 2
            candidates.append((score, s, p, info))

        candidates.sort(reverse=True)

        if candidates:
            _, slots, poles, info = candidates[0]
            return slots, poles
        return 36, 6

    def _recommend_lamination(self, spec: ProjectSpec) -> str:
        if spec.target_efficiency_pct >= 96:
            return "M190-27A"
        elif spec.target_efficiency_pct >= 93:
            return "M250-35A"
        return "M400-50A"

    def _recommend_magnet(self, spec: ProjectSpec) -> str:
        if spec.cooling in ("oil_spray", "oil_immersion"):
            return "NdFeB_42UH"
        elif spec.max_speed_rpm > 8000:
            return "NdFeB_42UH"
        return "NdFeB_35SH"


# ---------------------------------------------------------------------------
# File loading
# ---------------------------------------------------------------------------

def _parse_simple_yaml(content: str) -> dict:
    """Simple YAML key-value parser (no pyyaml dependency)."""
    data = {}
    for line in content.split("\n"):
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if ":" in line:
            key, _, val = line.partition(":")
            key = key.strip().lower().replace(" ", "_")
            val = val.strip().strip("\"'")
            # Try convert to number
            try:
                if "." in val:
                    val = float(val)
                else:
                    val = int(val)
            except ValueError:
                pass
            data[key] = val
    return data


def load_spec_from_file(filepath: str) -> ProjectSpec:
    """Load project spec from JSON, YAML, or TXT file."""
    ext = os.path.splitext(filepath)[1].lower()
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    if ext in (".yaml", ".yml"):
        data = _parse_simple_yaml(content)
    elif ext == ".json":
        data = json.loads(content)
    else:
        return parse_spec_from_text(content)

    # Map common key names
    mapping = {
        "project": "project_name", "project_name": "project_name", "name": "project_name",
        "motor_type": "motor_type", "type": "motor_type",
        "rated_power_kw": "rated_power_kw", "power_kw": "rated_power_kw", "power": "rated_power_kw",
        "rated_speed_rpm": "rated_speed_rpm", "speed_rpm": "rated_speed_rpm",
        "max_speed_rpm": "max_speed_rpm",
        "rated_torque_nm": "rated_torque_nm", "torque_nm": "rated_torque_nm",
        "peak_torque_nm": "peak_torque_nm",
        "dc_voltage_v": "dc_voltage_v", "voltage_v": "dc_voltage_v", "voltage": "dc_voltage_v",
        "max_current_arms": "max_current_arms",
        "target_efficiency_pct": "target_efficiency_pct", "efficiency": "target_efficiency_pct",
        "max_outer_diameter_mm": "max_outer_diameter_mm", "outer_dia": "max_outer_diameter_mm",
        "max_stack_length_mm": "max_stack_length_mm", "stack_length": "max_stack_length_mm",
        "cooling": "cooling",
        "application": "application", "app": "application",
        "max_current_density": "max_current_density",
    }

    kwargs = {}
    for src_key, dst_key in mapping.items():
        if src_key in data:
            kwargs[dst_key] = data[src_key]

    if "rated_power_kw" in kwargs and "max_current_arms" not in kwargs and "dc_voltage_v" in kwargs:
        kwargs["max_current_arms"] = round(kwargs["rated_power_kw"] * 1000 / kwargs["dc_voltage_v"] / 0.85, 1)

    return ProjectSpec(**kwargs)




# ---------------------------------------------------------------------------
# Excel loading
# ---------------------------------------------------------------------------

def _detect_spec_columns(headers: list) -> dict:
    """Auto-detect spec column names from Excel headers (Chinese + English)."""
    mapping = {}
    name_variants = {
        "project_name": ["project", "project_name", "name", "项目", "项目名称", "名称"],
        "motor_type": ["motor_type", "type", "电机类型", "类型"],
        "rated_power_kw": ["rated_power_kw", "power_kw", "power", "额定功率", "功率", "功率(kw)", "功率（kw）"],
        "rated_speed_rpm": ["rated_speed_rpm", "speed_rpm", "speed", "额定转速", "转速", "转速(rpm)", "转速（rpm）"],
        "max_speed_rpm": ["max_speed_rpm", "max_speed", "最高转速", "峰值转速"],
        "rated_torque_nm": ["rated_torque_nm", "torque_nm", "torque", "额定转矩", "转矩", "转矩(nm)", "转矩（nm）"],
        "peak_torque_nm": ["peak_torque_nm", "peak_torque", "峰值转矩", "最大转矩"],
        "dc_voltage_v": ["dc_voltage_v", "voltage_v", "voltage", "母线电压", "电压", "电压(v)", "电压（v）"],
        "max_current_arms": ["max_current_arms", "current_arms", "电流", "额定电流"],
        "target_efficiency_pct": ["target_efficiency_pct", "efficiency", "效率", "目标效率", "效率(%)", "效率（%）"],
        "max_outer_diameter_mm": ["max_outer_diameter_mm", "outer_dia", "外径", "最大外径", "外径(mm)", "外径（mm）"],
        "max_stack_length_mm": ["max_stack_length_mm", "stack_length", "叠长", "铁心长", "叠厚", "长度(mm)", "长度（mm）"],
        "cooling": ["cooling", "冷却方式", "冷却"],
        "application": ["application", "app", "应用", "用途"],
        "max_current_density": ["max_current_density", "电流密度", "电密"],
    }

    for idx, header in enumerate(headers):
        h = str(header).strip().lower()
        import re
        h = re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', h)
        best_field = None
        best_len = 0
        for field, variants in name_variants.items():
            norm_variants = [re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', v.lower()) for v in variants]
            for nv in norm_variants:
                # Match only if variant starts at beginning of header, or header in variant
                # This prevents "转速rpm" from matching "最高转速rpm"
                if (h.startswith(nv) or nv.startswith(h)) and len(nv) > best_len:
                    best_len = len(nv)
                    best_field = field
        if best_field:
            mapping[best_field] = idx
    return mapping


def load_spec_from_excel(filepath: str, row: int = 0) -> ProjectSpec:
    """Load project spec from an Excel file (.xlsx).

    Supports both single-row (row=0) and multi-row (specify which row).
    Auto-detects column headers in Chinese or English.

    Args:
        filepath: Path to .xlsx file.
        row: Row index (0-based, after header). Default first data row.

    Returns:
        ProjectSpec object.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Read headers from first row
    headers = [cell.value for cell in ws[1]]
    col_map = _detect_spec_columns(headers)

    # Read data from specified row
    data_row = row + 2  # Excel is 1-based, row 1 is header
    kwargs = {}
    for field, col_idx in col_map.items():
        value = ws.cell(row=data_row, column=col_idx + 1).value
        if value is not None:
            if field in ("project_name", "motor_type", "cooling", "application"):
                kwargs[field] = str(value).strip()
            else:
                try:
                    kwargs[field] = float(value)
                except (ValueError, TypeError):
                    kwargs[field] = str(value).strip()

    # Handle cooling normalization
    cooling_map = {
        "水冷": "water_jacket", "water": "water_jacket", "water_jacket": "water_jacket",
        "油冷": "oil_spray", "oil": "oil_spray", "oil_spray": "oil_spray",
        "风冷": "forced_air", "air": "forced_air", "forced_air": "forced_air",
        "自然冷却": "natural_convection", "natural": "natural_convection",
    }
    if "cooling" in kwargs:
        k = str(kwargs["cooling"]).lower().replace(" ", "_")
        kwargs["cooling"] = cooling_map.get(k, kwargs["cooling"])

    return ProjectSpec(**kwargs)


def load_specs_from_excel(filepath: str) -> list:
    """Load multiple project specs from an Excel file (one per row)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    specs = []
    for row_idx in range(2, ws.max_row + 1):
        try:
            spec = load_spec_from_excel(filepath, row=row_idx - 2)
            specs.append(spec)
        except Exception:
            continue
    return specs


def parse_spec_from_text(text: str) -> ProjectSpec:
    """Parse project specification from natural language text."""
    kwargs = {"project_name": "Parsed Project"}

    patterns = {
        "rated_power_kw": r"(\d+\.?\d*)\s*kW",
        "rated_speed_rpm": r"(\d+\.?\d*)\s*rpm",
        "max_speed_rpm": r"(\d+\.?\d*)\s*(?:最高|max).*?(\d+\.?\d*)\s*rpm",
        "dc_voltage_v": r"(\d+\.?\d*)\s*V",
        "target_efficiency_pct": r"(\d+\.?\d*)\s*%",
        "max_outer_diameter_mm": r"(?:外径|直径).*?(\d+\.?\d*)\s*mm",
        "max_stack_length_mm": r"(?:叠长|铁心长|长度).*?(\d+\.?\d*)\s*mm",
    }

    power_match = re.search(patterns["rated_power_kw"], text)
    if power_match:
        kwargs["rated_power_kw"] = float(power_match.group(1))

    speed_match = re.search(r"(\d+\.?\d*)\s*rpm", text)
    if speed_match:
        kwargs["rated_speed_rpm"] = float(speed_match.group(1))

    voltage_match = re.search(r"(\d+\.?\d*)\s*V", text)
    if voltage_match:
        kwargs["dc_voltage_v"] = float(voltage_match.group(1))

    eff_match = re.search(r"(\d+\.?\d*)\s*%", text)
    if eff_match:
        kwargs["target_efficiency_pct"] = float(eff_match.group(1))

    if "水冷" in text or "water" in text.lower():
        kwargs["cooling"] = "water_jacket"
    elif "油冷" in text or "oil" in text.lower():
        kwargs["cooling"] = "oil_spray"
    elif "风冷" in text or "air" in text.lower():
        kwargs["cooling"] = "forced_air"
    elif "自然" in text:
        kwargs["cooling"] = "natural_convection"

    if "IPM" in text.upper() or "内置" in text:
        kwargs["motor_type"] = "IPM"
    elif "SPM" in text.upper() or "表贴" in text:
        kwargs["motor_type"] = "SPM"
    elif "感应" in text or "IM" in text.upper():
        kwargs["motor_type"] = "IM"

    if "牵引" in text or "traction" in text.lower() or "EV" in text.upper():
        kwargs["application"] = "traction"
    elif "伺服" in text or "servo" in text.lower():
        kwargs["application"] = "servo"
    elif "直驱" in text or "direct" in text.lower():
        kwargs["application"] = "direct_drive"

    try:
        project_name = text.split("\n")[0].strip().rstrip("：:")
        if project_name and len(project_name) < 80:
            kwargs["project_name"] = project_name
    except Exception:
        pass

    return ProjectSpec(**kwargs)